import os
import PyPDF2
import re
import pandas as pd
from openpyxl import load_workbook
import shutil
import xml.etree.ElementTree as ET

DIST = 'COMPAGÁS'

class ExtratorFaturas:
    def __init__(self):
        self.regexes = {
            'cnpj': r'Gás\s?(\d+\.?\d+\.?\d+\/?\d+\-?\d+)\s?',  #19/09
            'valor_total': r'R\$\(?\)?\s?\:?\s?(\d+?.?\d+.?\d+\,\d{2})', #19/09 
            'volume_total': r'\s?(\d+?\.?\,?\d+\,?\.?\d+\,?.?\d+)\s?total\sm3?', #19/09
            'data_emissao': r'missão\s?(\d+\/\d+\/\d+)\s?', #19/09 (AJUSTAR ESTE REGEX)
            'data_inicio': r'consumo:?\s(\d+\/\d+\/\d+)', #19/09
            'data_fim': r'[A-a]\s(\d+\/\d+\/\d+)', #19/09             
            'numero_documento': r'(\d+)\s[E-e]sta', #19/09 (REGEX DE CORNO POR FALTA DE FATURA)
            'valor_icms': r'\s(\d+?\.?\,?\d+\,?\.?\d+\,?.?\d+)[T-t]ributo?', #19/09      
        }

    def extrair_informacoes(self, texto):
        informacoes = {}
        for chave, regex in self.regexes.items():
            match = re.search(regex, texto)
            if match:
                informacoes[chave] = match.group(1) if match.groups() else match.group(0)
        return informacoes

    def extrair_informacoes_xml(self, root):
        informacoes = {}
        for chave in self.regexes.keys():
            elemento = root.find(chave)
            if elemento is not None:
                informacoes[chave] = elemento.text
        return informacoes

def extrair_texto(caminho_do_pdf):
    texto = ''
    with open(caminho_do_pdf, 'rb') as arquivo:
        leitor_pdf = PyPDF2.PdfReader(arquivo)
        for pagina in leitor_pdf.pages:
            texto_pagina = pagina.extract_text()
            if texto_pagina:
                texto_pagina = texto_pagina.replace('\n', ' ')
                texto += texto_pagina + ' '
    
    if not texto:
        print(f"Erro ao extrair texto do PDF: {caminho_do_pdf}")
    else:
        print(f"Texto extraído do PDF {caminho_do_pdf}: {texto[:500]}...")  # Mostra os primeiros 500 caracteres do texto extraído
    return texto.strip()  # Remove espaços extras no início e no fim

def extrair_texto_xml(caminho_do_xml):
    try:
        tree = ET.parse(caminho_do_xml)
        root = tree.getroot()
        print(f"XML carregado com sucesso: {caminho_do_xml}")
        return root
    except Exception as e:
        print(f"Erro ao extrair texto do XML: {caminho_do_xml}, erro: {e}")
        return None

def registro_existe(df, cnpj, data_inicio, data_fim, valor_total):
    return not df[(df['CNPJ'] == cnpj) & (df['DATA INICIO'] == data_inicio) & (df['DATA FIM'] == data_fim) & (df['VALOR TOTAL'] == valor_total)].empty

def todos_campos_preenchidos(informacoes):
    campos_obrigatorios = ['cnpj', 'valor_total', 'volume_total', 'data_emissao', 'data_inicio', 'data_fim', 'numero_documento', 'valor_icms']
    for campo in campos_obrigatorios:
        if campo not in informacoes or not informacoes[campo]:
            print(f"Campo obrigatório '{campo}' está faltando ou vazio.")
            return False
    return True

def adicionar_na_planilha(informacoes, caminho_planilha, nome_arquivo):
    if not todos_campos_preenchidos(informacoes):
        print("Não foi possível adicionar à planilha devido a campos faltantes ou vazios.")
        return False

    try:
        df = pd.read_excel(caminho_planilha)
    except FileNotFoundError:
        print(f"O arquivo '{caminho_planilha}' não foi encontrado. Criando um novo.")
        df = pd.DataFrame(columns=['CNPJ', 'VALOR TOTAL', 'VOLUME TOTAL', 'DATA EMISSAO', 'DATA INICIO', 'DATA FIM', 'NUMERO FATURA', 'VALOR ICMS', 'DISTRIBUIDORA', 'NOME DO ARQUIVO'])
    
    cnpj = informacoes['cnpj']
    data_inicio = informacoes['data_inicio']
    data_fim = informacoes['data_fim']
    valor_total = pd.to_numeric(informacoes['valor_total'].replace('.', '').replace(',', '.'))
    volume_total = pd.to_numeric(informacoes['volume_total'].replace('.', '').replace(',', '.'))
    valor_icms = pd.to_numeric(informacoes['valor_icms'].replace('.', '').replace(',', '.'))

    if registro_existe(df, cnpj, data_inicio, data_fim, valor_total):
        print(f"Registro duplicado encontrado. Não será inserido.")
        return False 
    
    nova_linha = pd.DataFrame([{
        'CNPJ': cnpj,
        'VALOR TOTAL': valor_total,
        'VOLUME TOTAL': volume_total,
        'DATA EMISSAO': informacoes['data_emissao'],
        'DATA INICIO': data_inicio,
        'DATA FIM': data_fim,
        'NUMERO FATURA': informacoes['numero_documento'],
        'VALOR ICMS': valor_icms,
        'DISTRIBUIDORA': DIST,
        'NOME DO ARQUIVO': nome_arquivo
    }])
    df = pd.concat([df, nova_linha], ignore_index=True)
    df.to_excel(caminho_planilha, index=False)
    print("Dados adicionados com sucesso à planilha.")
    return True

def mover_arquivo(origem, destino):
    shutil.move(origem, destino)
    print(f"Arquivo movido para {destino}")

def verificar_linha_preenchida(caminho_planilha, informacoes):
    try:
        workbook = load_workbook(caminho_planilha)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Ignora o cabeçalho
            if (
                row[0] == informacoes.get('cnpj') and
                row[1] == informacoes.get('valor_total') and
                row[2] == informacoes.get('volume_total') and
                row[3] == informacoes.get('data_emissao') and
                row[4] == informacoes.get('data_inicio') and
                row[5] == informacoes.get('data_fim') and
                row[6] == informacoes.get('numero_documento') and
                row[7] == informacoes.get('valor_icms')

            ):
                if all(cell is not None and cell != '' for cell in row):
                    return True
                else:
                    return False
        return False  # Retorna False se a linha correspondente não for encontrada
    except Exception as e:
        print(f"Erro ao verificar a planilha: {e}")
        return False

def main(file_path, file, caminho_planilha):
    if file.lower().endswith('.pdf'):
        texto = extrair_texto(file)
        extrator = ExtratorFaturas()
        informacoes = extrator.extrair_informacoes(texto)
    elif file.lower().endswith('.xml'):
        root = extrair_texto_xml(file)
        if root is not None:
            extrator = ExtratorFaturas()
            informacoes = extrator.extrair_informacoes_xml(root)
        else:
            informacoes = {}
    else:
        print(f"Formato de arquivo não suportado: {file}")
        return

    if not any(informacoes.values()):
        print(f"Nenhuma informação extraída do arquivo: {file}")
        return

    nome_arquivo = os.path.basename(file)
    inserido = adicionar_na_planilha(informacoes, caminho_planilha, nome_arquivo)
    print(informacoes)

    if inserido:
        destino = os.path.join(diretorio_destino, nome_arquivo)
        mover_arquivo(file, destino)
    else:
        print('Arquivo não foi inserido na planilha devido a dados faltantes ou duplicados. Não será movido.')

# Exemplo de uso
file_path = r'G:\QUALIDADE\Códigos\Leitura de Faturas Gás\Códigos\Compagás\Faturas'
diretorio_destino = r'G:\QUALIDADE\Códigos\Leitura de Faturas Gás\Códigos\Compagás\Lidos'
caminho_planilha = r'G:\QUALIDADE\Códigos\Leitura de Faturas Gás\Códigos\00 Faturas Lidas\COMPAGÁS.xlsx'

for arquivo in os.listdir(file_path):
    if arquivo.lower().endswith('.pdf') or arquivo.lower().endswith('.xml'):
        arquivo_full = os.path.join(file_path, arquivo)
        arquivo = os.path.basename(arquivo)

        main(arquivo, arquivo_full, caminho_planilha)